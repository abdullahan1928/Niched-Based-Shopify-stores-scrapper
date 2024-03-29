import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment
import openpyxl
from bs4 import BeautifulSoup
import calc_time as ct
import excel_logic as ex
import time
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

ct.start_time

excel_file_to_be_saved = "camping_gear_final.xlsx"

dataframe = openpyxl.load_workbook("camping_gear.xlsx")
dataframe1 = dataframe.active

totalIndex = 1
notAddedToExcel = 0
instagramArray = []

session = requests.Session()
session = requests.Session()
retry = Retry(connect=3, backoff_factor=0.5)
adapter = HTTPAdapter(max_retries=retry)
session.mount('http://', adapter)
session.mount('https://', adapter)


for row in range(1, dataframe1.max_row):
    for col in dataframe1.iter_cols(1, 1):
        try:
            page = session.get(col[row].value)
            soup = BeautifulSoup(page.content, 'html.parser')
        except:
            time.sleep(10)
            continue

        try:
            brandName = soup.find(
                'meta', attrs={'property': 'og:site_name'})['content']
        except:
            brandName = 'None'

        instagram = soup.findAll('a', attrs={'href': True})

        for i in instagram:
            if 'instagram' in i['href']:
                instagram = i['href']
                break
            else:
                instagram = 'None'
                notAddedToExcel += 1

        if instagram != 'None' and instagram not in instagramArray:
            ex.ws.cell(row=row+1, column=1).value = brandName
            ex.ws.cell(row=row+1, column=2).value = col[row].value
            ex.ws.cell(row=row+1, column=3).value = instagram

            instagramArray.append(instagram)

            print(totalIndex, col[row].value, brandName, instagram)

        if totalIndex % 20 == 0:
            ex.wb.save(excel_file_to_be_saved)
            print("Saved")
            time.sleep(10)

        totalIndex += 1

print("Not added to excel: ", notAddedToExcel)

ex.wb.save(excel_file_to_be_saved)

ct.calc_total_time()
