from openpyxl import Workbook
from openpyxl.styles import Alignment

wb = Workbook()
ws = wb.active

# set column width


def set_column_widths():
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 30

# set column names


column_names = ['Brand Name', 'Website', 'Instagram']


def set_column_names():
    for index, column_name in enumerate(column_names):
        ws.cell(row=1, column=index+1).value = column_name
        # ws.cell(row=1, column=index+1).alignment = Alignment(horizontal='center')


set_column_widths()
set_column_names()
